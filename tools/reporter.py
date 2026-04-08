"""Generate Excel and text reports for suggested accruals and anomalies."""

import json
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Colour palette ───────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HIGH_FILL   = PatternFill("solid", fgColor="FF9999")
MED_FILL    = PatternFill("solid", fgColor="FFD966")
LOW_FILL    = PatternFill("solid", fgColor="C6EFCE")
ALT_FILL    = PatternFill("solid", fgColor="EBF0FA")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws, cols: list[str], row: int = 1):
    for col_idx, title in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def generate_excel_report(
    suggestions: list[dict],
    anomalies: list[dict],
    period_trends: list[dict],
    target_period: str,
    output_dir: str = ".",
) -> str:
    """
    Write a multi-sheet Excel workbook:
      1. Accruals to Post   — suggested accruals for target_period
      2. Anomalies Detected — irregularity report
      3. Period Trends      — historical summary
    Returns the output file path.
    """
    wb = Workbook()

    # ── Sheet 1: Accruals to Post ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Accruals to Post"

    accrual_cols = [
        "Company Code", "Company Name", "GL Account", "GL Description",
        "Vendor Number", "Vendor Name", "Suggested Amount (USD)",
        "Basis", "Confidence", "Historical Mean (USD)", "Historical Std (USD)",
        "Periods Observed",
    ]
    _header_row(ws1, accrual_cols)

    for r_idx, s in enumerate(suggestions, start=2):
        row_data = [
            s["company_code"], s["company_name"], s["gl_account"], s["gl_description"],
            s["vendor_number"], s["vendor_name"],
            round(s["suggested_amount_usd"], 2),
            s["basis"], s["confidence"],
            round(s["historical_mean"], 2), round(s["historical_std"], 2),
            s["periods_observed"],
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            # Colour by confidence
            conf = s["confidence"]
            if conf == "HIGH":
                cell.fill = LOW_FILL
            elif conf == "MEDIUM":
                cell.fill = ALT_FILL
            else:
                cell.fill = MED_FILL
        # Format amount column
        ws1.cell(row=r_idx, column=7).number_format = '#,##0.00'
        ws1.cell(row=r_idx, column=10).number_format = '#,##0.00'
        ws1.cell(row=r_idx, column=11).number_format = '#,##0.00'

    # Total row
    total_row = len(suggestions) + 2
    ws1.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)
    total_cell = ws1.cell(row=total_row, column=7,
                          value=sum(s["suggested_amount_usd"] for s in suggestions))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'

    _auto_width(ws1)
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Anomalies Detected ───────────────────────────────────────────
    ws2 = wb.create_sheet("Anomalies Detected")
    anomaly_cols = [
        "Sl.No.", "Company Code", "Company Name", "GL Account", "GL Description",
        "Vendor Number", "Vendor Name", "Amount (USD)", "Period",
        "Severity", "Anomaly Type", "Detail",
    ]
    _header_row(ws2, anomaly_cols)

    for r_idx, a in enumerate(anomalies, start=2):
        row_data = [
            a["sl_no"], a["company_code"], a["company_name"],
            a["gl_account"], a["gl_description"],
            a["vendor_number"], a["vendor_name"],
            round(a["amount_usd"], 2), a["period"],
            a["severity"], a["anomaly_type"], a["detail"],
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            sev = a["severity"]
            cell.fill = HIGH_FILL if sev == "HIGH" else MED_FILL if sev == "MEDIUM" else LOW_FILL
        ws2.cell(row=r_idx, column=8).number_format = '#,##0.00'

    _auto_width(ws2)
    ws2.freeze_panes = "A2"

    # ── Sheet 3: Period Trends ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Period Trends")
    trend_cols = ["Period", "# Entries", "Total Amount (USD)", "Average Amount (USD)"]
    _header_row(ws3, trend_cols)

    for r_idx, t in enumerate(period_trends, start=2):
        row_data = [
            t["period"],
            int(t["count"]),
            round(float(t["total"]), 2),
            round(float(t["average"]), 2),
        ]
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            cell.fill = ALT_FILL if r_idx % 2 == 0 else PatternFill()
        ws3.cell(row=r_idx, column=3).number_format = '#,##0.00'
        ws3.cell(row=r_idx, column=4).number_format = '#,##0.00'

    _auto_width(ws3)

    # ── Save ──────────────────────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Accrual_Report_{target_period.replace('/', '-')}_{ts}.xlsx"
    output_path = Path(output_dir) / filename
    wb.save(output_path)
    return str(output_path)


def generate_json_report(
    suggestions: list[dict],
    anomalies: list[dict],
    agent_analysis: str,
    target_period: str,
    output_dir: str = ".",
) -> str:
    """Write a machine-readable JSON report."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Accrual_Report_{target_period.replace('/', '-')}_{ts}.json"
    output_path = Path(output_dir) / filename

    report = {
        "generated_at": datetime.now().isoformat(),
        "target_period": target_period,
        "summary": {
            "total_suggested_entries": len(suggestions),
            "total_suggested_amount_usd": round(sum(s["suggested_amount_usd"] for s in suggestions), 2),
            "anomalies_detected": len(anomalies),
            "high_severity": sum(1 for a in anomalies if a["severity"] == "HIGH"),
            "medium_severity": sum(1 for a in anomalies if a["severity"] == "MEDIUM"),
            "low_severity": sum(1 for a in anomalies if a["severity"] == "LOW"),
        },
        "agent_analysis": agent_analysis,
        "suggested_accruals": suggestions,
        "anomalies": anomalies,
    }

    with open(output_path, "w") as f:
        json.dump(report, f, indent=2, default=str)

    return str(output_path)
